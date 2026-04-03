"""
main.py — FastAPI application for the CAT Modeling Data Pipeline.

Endpoints:
  POST   /upload                    Upload CSV/XLSX, create session
  GET    /suggest-columns/{id}      Get column mapping suggestions
  POST   /confirm-columns/{id}      Confirm column mapping
  POST   /geocode/{id}              Run geocoding
  POST   /map-codes/{id}            Run 4-stage code mapping
  POST   /normalize/{id}            Run normalization
  GET    /review/{id}               Get accumulated flags
  POST   /correct/{id}              Apply corrections
  GET    /download/{id}             Download output XLSX or CSV
  GET    /session/{id}              Session info
  GET    /sessions                  List all active sessions (debug)
  DELETE /session/{id}              Delete session
"""
import io
import json
import logging
import os
import pathlib
from contextlib import asynccontextmanager
from typing import Any, Dict, List, Optional

import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

# ── Load env early ─────────────────────────────────────────────────────────────
load_dotenv()

# ── Local imports ──────────────────────────────────────────────────────────────
import session as session_store
import code_mapper
import geocoder
import mapping_memory
from column_mapper import suggest_columns, validate_required_fields
from normalizer import normalize_all_rows
from output_builder import build_xlsx, build_csv
from rules import BusinessRulesConfig
from models import (
    UploadResponse, SuggestColumnsResponse, ColumnSuggestion,
    ConfirmColumnsRequest, ConfirmColumnsResponse,
    GeocodeResponse, MapCodesResponse, NormalizeResponse,
    ReviewResponse, FlagEntry, CorrectRequest, CorrectResponse,
    SessionInfoResponse,
)

load_dotenv()
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s [%(name)s] %(levelname)s — %(message)s",
)
logger = logging.getLogger("main")

# ── Lifespan: pre-build TF-IDF and start TTL cleanup ──────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Starting CAT pipeline — loading TF-IDF indexes…")
    code_mapper.build_tfidf_indexes()
    geocoder.load_reference_data()
    session_store.start_ttl_cleanup()
    logger.info("Startup complete.")
    yield
    logger.info("Shutting down.")


# ── App ────────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="CAT Modeling Data Pipeline",
    description="Upload, map, geocode, classify, and normalise property exposure data for AIR/RMS CAT models.",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_session_or_404(upload_id: str) -> dict:
    try:
        return session_store.require_session(upload_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Session '{upload_id}' not found or expired.")


def _require_stage(session: dict, stage: str, endpoint: str) -> None:
    if not session.get("stages_complete", {}).get(stage):
        raise HTTPException(
            status_code=422,
            detail=f"Stage '{stage}' must be completed before calling '{endpoint}'.",
        )


def _load_iso4217() -> set:
    p = pathlib.Path(__file__).parent / "reference" / "iso4217_currency.json"
    if p.exists():
        data = json.loads(p.read_text(encoding="utf-8"))
        return set(data.keys())
    return set()


_VALID_CURRENCIES = _load_iso4217()


def _enrich_excel_formats(content: bytes, df: pd.DataFrame) -> pd.DataFrame:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        col_currencies = {}
        for col_idx in range(1, ws.max_column + 1):
            if col_idx - 1 >= len(df.columns): break
            col_name = df.columns[col_idx - 1]
            for r in range(2, min(ws.max_row + 1, 10)):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    fmt = str(cell.number_format or "")
                    for sym in ["$", "€", "£", "¥", "₹"]:
                        if sym in fmt:
                            col_currencies[col_name] = sym
                            break
                    if col_name in col_currencies:
                        break
        for col, sym in col_currencies.items():
            df[col] = df[col].apply(lambda v: f"{sym}{v}" if pd.notnull(v) and str(v).strip() != "" else v)
    except Exception as e:
        logger.warning(f"Could not extract Excel currency formats: {e}")
    return df


# ── Step 1: Upload ─────────────────────────────────────────────────────────────

@app.post("/upload", response_model=UploadResponse, tags=["Pipeline"])
async def upload(
    file: UploadFile = File(...),
    target_format: str = Query("AIR", regex="^(AIR|RMS)$"),
    rules_config: Optional[str] = Form(None),
):
    """
    Upload a CSV or XLSX file and create a processing session.
    No row limit — all rows are ingested.
    """
    content = await file.read()
    fname = (file.filename or "").lower()

    # Parse rules config
    try:
        rules_dict = json.loads(rules_config) if rules_config and rules_config.strip() else {}
        rules = BusinessRulesConfig(**rules_dict)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Invalid rules_config: {exc}")

    # Parse file
    try:
        if fname.endswith(".csv"):
            try:
                df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False)
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, encoding="latin-1")
        elif fname.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False, sheet_name=0)
            df = _enrich_excel_formats(content, df)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload CSV or XLSX.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"File parsing error: {exc}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no data rows.")

    # Deduplicate column names
    seen: Dict[str, int] = {}
    new_cols = []
    for col in df.columns:
        col = str(col).strip()
        if col in seen:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 1
            new_cols.append(col)
    df.columns = new_cols

    # Clean values: strip whitespace, replace empty string with None
    df = df.map(lambda v: v.strip() if isinstance(v, str) and v.strip() != "" else None)

    headers = list(df.columns)
    raw_rows = df.to_dict(orient="records")
    sample = raw_rows[:5]

    upload_id = session_store.create_session({
        "target_format": target_format,
        "rules_config": rules.model_dump(),
        "raw_rows": raw_rows,
        "headers": headers,
        "sample": sample,
        "column_map": {},
        "unmapped_cols": [],
        "geo_rows": [],
        "code_map": {},
        "final_rows": [],
    })

    logger.info(f"Session {upload_id}: uploaded {len(raw_rows)} rows, format={target_format}")
    return UploadResponse(
        upload_id=upload_id,
        row_count=len(raw_rows),
        headers=headers,
        sample=sample,
        target_format=target_format,
    )


# ── Step 2a: Suggest columns ───────────────────────────────────────────────────

@app.get("/suggest-columns/{upload_id}", response_model=SuggestColumnsResponse, tags=["Pipeline"])
def suggest_columns_endpoint(upload_id: str):
    """Stateless: analyse column names and return ranked mapping suggestions."""
    session = _get_session_or_404(upload_id)
    raw_rows = session["raw_rows"]
    headers = session["headers"]

    # Build sample values dict (first 3 non-null per column)
    sample_values: Dict[str, List[Any]] = {}
    for col in headers:
        vals = [r[col] for r in raw_rows[:20] if r.get(col) is not None][:3]
        sample_values[col] = vals

    result = suggest_columns(
        source_columns=headers,
        sample_values=sample_values,
        target_format=session["target_format"],
        fuzzy_threshold=session["rules_config"].get("fuzzy_llm_fallback_threshold", 72),
        cutoff=session["rules_config"].get("fuzzy_score_cutoff", 50),
    )

    suggestions_typed = {
        col: [ColumnSuggestion(**s) for s in sug_list]
        for col, sug_list in result["suggestions"].items()
    }
    return SuggestColumnsResponse(
        suggestions=suggestions_typed,
        unmapped=result["unmapped"],
        memory_count=result.get("memory_count", 0),
    )


# ── Step 2b: Confirm columns ───────────────────────────────────────────────────

@app.post("/confirm-columns/{upload_id}", response_model=ConfirmColumnsResponse, tags=["Pipeline"])
def confirm_columns(upload_id: str, body: ConfirmColumnsRequest):
    """Confirm and persist the column mapping. Validates required fields and 1:1 uniqueness."""
    session = _get_session_or_404(upload_id)

    column_map = body.column_map

    # ── Enforce 1:1 mapping: one source → one canonical, one canonical ← one source ──
    # Collect which source columns map to each canonical (ignoring None/unmapped).
    canonical_to_sources: Dict[str, List[str]] = {}
    for src_col, canonical in column_map.items():
        if canonical is None:
            continue
        canonical_to_sources.setdefault(canonical, []).append(src_col)

    # Find violations: any canonical claimed by more than one source column
    duplicate_violations = {
        canonical: sources
        for canonical, sources in canonical_to_sources.items()
        if len(sources) > 1
    }

    if duplicate_violations:
        details = "; ".join(
            f"'{canonical}' claimed by: {sources}"
            for canonical, sources in duplicate_violations.items()
        )
        raise HTTPException(
            status_code=422,
            detail=f"Mapping violation — each canonical field must be mapped by exactly one source column. "
                   f"Duplicates found: {details}",
        )

    unmapped = [k for k, v in column_map.items() if v is None]
    mapped_count = len(column_map) - len(unmapped)

    missing_required = validate_required_fields(column_map, session["target_format"])
    warnings = [f"Required field '{f}' is not mapped." for f in missing_required]

    session_store.update_session(upload_id, {
        "column_map": column_map,
        "unmapped_cols": unmapped,
    })
    session_store.session_mark_stage(upload_id, "column_map")

    # ── Learn from this confirmed mapping ───────────────────────────────────────────
    try:
        mapping_memory.record_confirmed(column_map, session["target_format"])
    except Exception as exc:
        logger.warning(f"mapping_memory.record_confirmed failed (non-fatal): {exc}")

    logger.info(f"Session {upload_id}: column map confirmed, {mapped_count} mapped, {len(unmapped)} unmapped")
    return ConfirmColumnsResponse(
        upload_id=upload_id,
        mapped_count=mapped_count,
        unmapped_cols=unmapped,
        missing_required=missing_required,
        warnings=warnings,
    )


# ── Step 3: Geocode ────────────────────────────────────────────────────────────

@app.post("/geocode/{upload_id}", response_model=GeocodeResponse, tags=["Pipeline"])
def geocode_endpoint(upload_id: str):
    """Geocode all rows using Geoapify. Skips rows with pre-existing coordinates."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "column_map", "/geocode")

    column_map = session["column_map"]
    rules_config = BusinessRulesConfig(**session["rules_config"])

    # Remap raw rows using confirmed column_map
    raw_rows = session["raw_rows"]
    remapped = _apply_column_map(raw_rows, column_map)

    geocoded_count = provided_count = failed_count = 0
    new_flags: List[dict] = []
    geo_rows: List[dict] = []

    for idx, row in enumerate(remapped):
        geo_fields = geocoder.process_row_geocoding(
            row, column_map,
            target_format=session.get("target_format", "AIR"),
        )
        row.update(geo_fields)
        geo_rows.append(row)

        status = geo_fields.get("GeocodingStatus", "")
        source = geo_fields.get("Geosource", "")

        if source == "Provided":
            provided_count += 1
        elif status == "OK":
            geocoded_count += 1
            # State code validation flag
            if geo_fields.get("StateCodeValidation") == "UNRECOGNIZED":
                new_flags.append({
                    "row_index": idx,
                    "field": "Statecode_Final",
                    "issue": "unrecognized_state_code",
                    "current_value": geo_fields.get("Statecode_Final"),
                    "confidence": None,
                    "alternatives": [],
                    "message": f"State code '{geo_fields.get('Statecode_Final')}' not found in ISO 3166 reference",
                })
        else:
            failed_count += 1
            new_flags.append({
                "row_index": idx,
                "field": "GeocodingStatus",
                "issue": "geocoding_failed",
                "current_value": status,
                "confidence": None,
                "alternatives": [],
                "message": f"Geocoding failed for row {idx}: {status}",
            })

    session_store.update_session(upload_id, {"geo_rows": geo_rows})
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "geocoding")

    logger.info(f"Session {upload_id}: geocoded={geocoded_count}, provided={provided_count}, failed={failed_count}")
    return GeocodeResponse(
        upload_id=upload_id,
        total_rows=len(raw_rows),
        geocoded=geocoded_count,
        provided=provided_count,
        failed=failed_count,
        flags_added=len(new_flags),
    )


# ── Step 4: Map codes ──────────────────────────────────────────────────────────

@app.post("/map-codes/{upload_id}", response_model=MapCodesResponse, tags=["Pipeline"])
def map_codes_endpoint(upload_id: str):
    """
    Run 4-stage code mapping (deterministic → Gemini LLM → TF-IDF → default)
    for both occupancy and construction codes.
    """
    session = _get_session_or_404(upload_id)
    _require_stage(session, "geocoding", "/map-codes")

    geo_rows = session["geo_rows"]
    column_map = session["column_map"]
    rules_config = BusinessRulesConfig(**session["rules_config"])
    target = session["target_format"]

    # Identify schema columns
    occ_scheme_col, occ_value_col = _find_code_columns(column_map, "occupancy", target)
    const_scheme_col, const_value_col = _find_code_columns(column_map, "construction", target)

    target_occ_scheme = "OccupancyCodeType" if target == "AIR" else "OCCSCHEME"
    target_const_scheme = "ConstructionCodeType" if target == "AIR" else "BLDGSCHEME"

    new_flags: List[dict] = []
    enriched_rows = []
    code_map: Dict[str, Dict] = {}

    # ── Occupancy mapping ────────────────────────────────────────────────────
    if occ_value_col:
        occ_items = code_mapper.extract_unique_pairs(geo_rows, occ_scheme_col, occ_value_col)
        occ_results = code_mapper.map_codes(occ_items, target, "occupancy", rules_config)
        for item in occ_items:
            key = code_mapper.build_row_key(item["scheme"], item["value"])
            result = occ_results.get(str(item["index"]), {})
            if result:
                code_map[f"occ|{key}"] = result

    # ── Construction mapping ─────────────────────────────────────────────────
    if const_value_col:
        const_items = code_mapper.extract_unique_pairs(geo_rows, const_scheme_col, const_value_col)
        const_results = code_mapper.map_codes(const_items, target, "construction", rules_config)
        for item in const_items:
            key = code_mapper.build_row_key(item["scheme"], item["value"])
            result = const_results.get(str(item["index"]), {})
            if result:
                code_map[f"const|{key}"] = result

    # ── Enrich geo_rows with code results ────────────────────────────────────
    occ_by_method: Dict[str, int] = {}
    const_by_method: Dict[str, int] = {}

    enriched_rows = []
    for idx, row in enumerate(geo_rows):
        row = dict(row)

        # Apply occupancy
        if occ_value_col:
            scheme = str(row.get(occ_scheme_col) or "").strip()
            value = str(row.get(occ_value_col) or "").strip()
            key = f"occ|{code_mapper.build_row_key(scheme, value)}"
            result = code_map.get(key, {})
            if result:
                # ── Write mapped integer code to the canonical column the output builder reads ──
                row[occ_value_col] = result["code"]      # e.g. row["OccupancyCode"] = 302
                # Write scheme label; default ATC for RMS, AIR for AIR
                if not row.get(target_occ_scheme):
                    row[target_occ_scheme] = "ATC" if target == "RMS" else "AIR"
                row["Occupancy_Code"]        = result["code"]
                row["Occupancy_Description"] = result["description"]
                row["Occupancy_Confidence"]  = result["confidence"]
                row["Occupancy_Method"]      = result["method"]
                row["Occupancy_Original"]    = result["original"]
                occ_by_method[result["method"]] = occ_by_method.get(result["method"], 0) + 1
                if result["confidence"] < rules_config.occ_confidence_threshold:
                    new_flags.append({
                        "row_index": idx, "field": occ_value_col,
                        "issue": "low_confidence",
                        "current_value": result["code"],
                        "confidence": result["confidence"],
                        "alternatives": result.get("alternatives", []),
                        "message": (f"Occupancy '{value}' mapped to {result['code']} "
                                    f"({result['description']}) with low confidence {result['confidence']:.2f}"),
                    })

        # Apply construction
        if const_value_col:
            scheme = str(row.get(const_scheme_col) or "").strip()
            value = str(row.get(const_value_col) or "").strip()
            key = f"const|{code_mapper.build_row_key(scheme, value)}"
            result = code_map.get(key, {})
            if result:
                # ── Write mapped code to the canonical column ───────────────
                row[const_value_col] = result["code"]    # e.g. BLDGCLASS = "1"
                # scheme_override: if the lookup determined a different scheme
                # (e.g. Non-Combustible → FIRE instead of RMS), write it now.
                scheme_override = result.get("scheme_override")
                # scheme_override values ("RMS", "FIRE") are RMS EDM BLDGSCHEME
                # vocabulary — they must ONLY be applied when target is RMS.
                # For AIR targets, ConstructionCodeType usually is "AIR".
                # However, if the mapper detected an "ISF" classification, preserve it.
                if scheme_override == "ISF":
                    row[target_const_scheme] = "ISF"
                elif target == "RMS" and scheme_override:
                    row[target_const_scheme] = scheme_override
                elif not row.get(target_const_scheme):
                    row[target_const_scheme] = target
                # Preserve metadata in internal audit fields
                row["Construction_Code"]        = result["code"]
                row["Construction_Description"] = result["description"]
                row["Construction_Confidence"]  = result["confidence"]
                row["Construction_Method"]      = result["method"]
                row["Construction_Original"]    = result["original"]
                row["Construction_Scheme"]      = row.get(target_const_scheme, target)
                const_by_method[result["method"]] = const_by_method.get(result["method"], 0) + 1
                if result["confidence"] < rules_config.const_confidence_threshold:
                    new_flags.append({
                        "row_index": idx, "field": const_value_col,
                        "issue": "low_confidence",
                        "current_value": result["code"],
                        "confidence": result["confidence"],
                        "alternatives": result.get("alternatives", []),
                        "message": (f"Construction '{value}' mapped to {result['code']} "
                                    f"({result['description']}) with low confidence {result['confidence']:.2f}"),
                    })

        enriched_rows.append(row)

    session_store.update_session(upload_id, {
        "code_map": code_map,
        "final_rows": enriched_rows,
    })
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "code_mapping")

    logger.info(f"Session {upload_id}: code mapping complete. "
                f"occ_methods={occ_by_method}, const_methods={const_by_method}")

    return MapCodesResponse(
        upload_id=upload_id,
        unique_occ_pairs=len([k for k in code_map if k.startswith("occ|")]),
        unique_const_pairs=len([k for k in code_map if k.startswith("const|")]),
        occ_by_method=occ_by_method,
        const_by_method=const_by_method,
        flags_added=len(new_flags),
    )


# ── Step 5: Normalize ──────────────────────────────────────────────────────────

@app.post("/normalize/{upload_id}", response_model=NormalizeResponse, tags=["Pipeline"])
def normalize_endpoint(upload_id: str):
    """Run all normalization (year, stories, area, values, currency, modifiers)."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "code_mapping", "/normalize")

    rules_config = BusinessRulesConfig(**session["rules_config"])
    final_rows = session["final_rows"] or session["geo_rows"]

    # Reload valid currencies
    valid_currencies = _VALID_CURRENCIES

    target_format = session.get("target_format", "AIR")
    normalized, new_flags = normalize_all_rows(final_rows, rules_config, valid_currencies,
                                                target_format=target_format)

    # Apply global line of business if provided
    lob_col = "LineOfBusiness" if target_format == "AIR" else ""
    if lob_col and rules_config.line_of_business:
        for r in normalized:
            if not r.get(lob_col):
                r[lob_col] = rules_config.line_of_business

    session_store.update_session(upload_id, {"final_rows": normalized})
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "normalization")

    # Summarise what was done
    summary = {
        "year_flags": sum(1 for f in new_flags if "year" in f.get("issue", "")),
        "story_flags": sum(1 for f in new_flags if "story" in f.get("issue", "") or "stories" in f.get("issue", "")),
        "area_flags": sum(1 for f in new_flags if "area" in f.get("issue", "")),
        "value_flags": sum(1 for f in new_flags if "value" in f.get("issue", "")),
        "currency_flags": sum(1 for f in new_flags if "currency" in f.get("issue", "")),
    }

    logger.info(f"Session {upload_id}: normalization complete, {len(new_flags)} new flags")
    return NormalizeResponse(
        upload_id=upload_id,
        total_rows=len(normalized),
        flags_added=len(new_flags),
        normalization_summary=summary,
    )


# ── Review & Corrections ───────────────────────────────────────────────────────

@app.get("/review/{upload_id}", response_model=ReviewResponse, tags=["Review"])
def review(upload_id: str):
    """Return all accumulated flags for the session."""
    session = _get_session_or_404(upload_id)
    flags = [FlagEntry(**f) for f in session.get("flags", [])]
    return ReviewResponse(
        upload_id=upload_id,
        flags=flags,
        stages_complete=session.get("stages_complete", {}),
    )


@app.post("/correct/{upload_id}", response_model=CorrectResponse, tags=["Review"])
def correct(upload_id: str, body: CorrectRequest):
    """
    Apply manual corrections to final_rows.
    Each correction:
      • Updates final_rows[row_index][field]
      • Removes the matching flag
      • If occupancy/construction code, updates code_map for that key
    """
    session = _get_session_or_404(upload_id)
    final_rows = session.get("final_rows", [])
    code_map = session.get("code_map", {})
    rules_config = BusinessRulesConfig(**session["rules_config"])
    applied = 0
    flags_removed = 0

    for correction in body.corrections:
        idx = correction.row_index
        field = correction.field
        value = correction.new_value

        if idx < 0 or idx >= len(final_rows):
            continue

        final_rows[idx][field] = value

        # If this is a code override, update code_map so other rows with same source benefit
        if field in ("Occupancy_Code", "Construction_Code"):
            prefix = "occ|" if field == "Occupancy_Code" else "const|"
            scheme = final_rows[idx].get("OccupancyScheme" if field == "Occupancy_Code" else "ConstructionScheme", "")
            original = final_rows[idx].get(
                "Occupancy_Original" if field == "Occupancy_Code" else "Construction_Original", "")
            key = f"{prefix}{code_mapper.build_row_key(scheme, original)}"
            if key in code_map:
                code_map[key]["code"] = value
                code_map[key]["method"] = "user_override"
                code_map[key]["confidence"] = 1.0
                # Propagate to all matching rows
                for r in final_rows:
                    r_orig = r.get("Occupancy_Original" if field == "Occupancy_Code" else "Construction_Original", "")
                    r_scheme = r.get("OccupancyScheme" if field == "Occupancy_Code" else "ConstructionScheme", "")
                    if r_orig == original and r_scheme == scheme:
                        r[field] = value
                        r[field.replace("_Code", "_Method")] = "user_override"
                        r[field.replace("_Code", "_Confidence")] = 1.0

        # Remove flag
        session_store.remove_flag(upload_id, idx, field)
        flags_removed += 1
        applied += 1

    session_store.update_session(upload_id, {
        "final_rows": final_rows,
        "code_map": code_map,
    })

    return CorrectResponse(applied=applied, flags_removed=flags_removed)


# ── Download ───────────────────────────────────────────────────────────────────

@app.get("/download/{upload_id}", tags=["Output"])
def download(
    upload_id: str,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
):
    """Download the processed output as XLSX or CSV."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/download")

    final_rows = session.get("final_rows", [])
    unmapped_cols = session.get("unmapped_cols", [])
    flags = session.get("flags", [])
    target = session.get("target_format", "AIR")
    short_id = upload_id[:8]

    if format == "csv":
        buf = build_csv(final_rows, unmapped_cols, target)
        filename = f"cat_output_{short_id}.csv"
        media_type = "text/csv"
    else:
        buf = build_xlsx(final_rows, unmapped_cols, flags, target, upload_id)
        filename = f"cat_output_{short_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    logger.info(f"Session {upload_id}: download requested ({format}), {len(final_rows)} rows")
    return StreamingResponse(
        buf,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Session management ─────────────────────────────────────────────────────────

@app.get("/session/{upload_id}", response_model=SessionInfoResponse, tags=["Session"])
def session_info(upload_id: str):
    session = _get_session_or_404(upload_id)
    return SessionInfoResponse(
        upload_id=upload_id,
        created_at=session["created_at"],
        target_format=session.get("target_format", "AIR"),
        row_count=len(session.get("raw_rows", [])),
        stages_complete=session.get("stages_complete", {}),
        flag_count=len(session.get("flags", [])),
        rules_config=session.get("rules_config", {}),
    )


@app.get("/sessions", tags=["Session"])
def list_sessions():
    return {"sessions": session_store.list_sessions()}


@app.delete("/session/{upload_id}", tags=["Session"])
def delete_session(upload_id: str):
    if session_store.delete_session(upload_id):
        return {"deleted": upload_id}
    raise HTTPException(status_code=404, detail="Session not found.")


# ── Pipeline Diff ──────────────────────────────────────────────────────────────────

@app.get("/session-diff/{upload_id}", tags=["Pipeline"])
def session_diff(upload_id: str, step: str = Query(..., regex="^(geocode|map-codes|normalize)$")):
    """
    Return before/after table data for a specific pipeline step, capped at 100 rows.
    Includes all columns that appear in the final Excel output.
    Also returns a `pairs` list: [{before_col, after_col, label}] so the UI can
    render old → new columns adjacently.
    """
    session = _get_session_or_404(upload_id)
    target = session.get("target_format", "AIR")
    _require_stage(session, "column_map", f"/session-diff (step={step})")

    column_map = session.get("column_map", {})
    raw_rows = session.get("raw_rows", [])

    # Reverse map: canonical → source column name
    canonical_to_src: Dict[str, str] = {}
    for src, can in column_map.items():
        if can and can not in canonical_to_src:
            canonical_to_src[can] = src

    def get_source_cols(canonicals: set) -> List[str]:
        return [src for src, can in column_map.items() if can in canonicals]

    # pairs = [{"label": str, "before": src_col | None, "after": canonical_col | None}]
    pairs: List[Dict] = []
    after_rows: List[Dict] = []

    if step == "geocode":
        _require_stage(session, "geocoding", f"/session-diff (step={step})")
        after_rows = session.get("geo_rows", [])

        # Detect if user provided a single FullAddress column instead of separate fields
        full_addr_src = canonical_to_src.get("FullAddress")  # source col name, or None
        full_address_mode = bool(full_addr_src)

        if target == "AIR":
            addr_fields = [
                ("Street",      "Street"),
                ("City",        "City"),
                ("Area",        "Area"),
                ("PostalCode",  "PostalCode"),
                ("CountryISO",  "CountryISO"),
                ("Latitude",    "Latitude"),
                ("Longitude",   "Longitude"),
            ]
        else:
            addr_fields = [
                ("STREETNAME",  "STREETNAME"),
                ("CITY",        "CITY"),
                ("STATECODE",   "STATECODE"),
                ("POSTALCODE",  "POSTALCODE"),
                ("CNTRYCODE",   "CNTRYCODE"),
                ("Latitude",    "Latitude"),
                ("Longitude",   "Longitude"),
            ]

        if full_address_mode:
            # Single input (FullAddress) → many extracted outputs.
            # Each pair's "before" is the full-address source col so the UI can show
            # the raw string alongside each extracted component.
            for canonical, after_col in addr_fields:
                pairs.append({
                    "label": after_col,
                    "before": full_addr_src,    # always the same source column
                    "after":  after_col,
                    "before_is_full_address": True,  # hint for the UI
                })
        else:
            # Normal field-by-field mapping
            for canonical, after_col in addr_fields:
                src = canonical_to_src.get(canonical)
                pairs.append({"label": canonical, "before": src, "after": after_col})

        # Geocoding-only outputs (no source equivalent)
        pairs.append({"label": "GeocodingStatus", "before": None, "after": "GeocodingStatus"})
        pairs.append({"label": "Geosource",       "before": None, "after": "Geosource"})

    elif step == "map-codes":
        _require_stage(session, "code_mapping", f"/session-diff (step={step})")
        after_rows = session.get("final_rows", session.get("geo_rows", []))

        if target == "AIR":
            code_pairs = [
                ("OccupancyCodeType",    "OccupancyCodeType",    "Occupancy_Code",         "Occ Code"),
                ("OccupancyCode",        "OccupancyCode",        "Occupancy_Description",  "Occ Description"),
                (None,                   None,                   "Occupancy_Method",       "Occ Method"),
                ("ConstructionCodeType", "ConstructionCodeType", "Construction_Code",      "Const Code"),
                ("ConstructionCode",     "ConstructionCode",     "Construction_Description","Const Description"),
                (None,                   None,                   "Construction_Method",    "Const Method"),
            ]
        else:
            code_pairs = [
                ("OCCSCHEME",  "OCCSCHEME",  "Occupancy_Code",         "Occ Code"),
                ("OCCTYPE",    "OCCTYPE",    "Occupancy_Description",  "Occ Description"),
                (None,         None,         "Occupancy_Method",       "Occ Method"),
                ("BLDGSCHEME", "BLDGSCHEME", "Construction_Code",      "Const Code"),
                ("BLDGCLASS",  "BLDGCLASS",  "Construction_Description","Const Description"),
                (None,         None,         "Construction_Method",    "Const Method"),
            ]

        for canonical, _, after_col, label in code_pairs:
            src = canonical_to_src.get(canonical) if canonical else None
            pairs.append({"label": label, "before": src, "after": after_col})

    elif step == "normalize":
        _require_stage(session, "normalization", f"/session-diff (step={step})")
        after_rows = session.get("final_rows", [])

        if target == "AIR":
            norm_pairs = [
                ("YearBuilt",        "YearBuilt"),
                ("YearRetrofitted",  "YearRetrofitted"),
                ("NumberOfStories",  "NumberOfStories"),
                ("RiskCount",        "RiskCount"),
                ("GrossArea",        "GrossArea"),
                ("BuildingValue",    "BuildingValue"),
                ("ContentsValue",    "ContentsValue"),
                ("TimeElementValue", "TimeElementValue"),
                ("Currency",         "Currency"),
                ("LineOfBusiness",   "LineOfBusiness"),
                ("SprinklerSystem",  "SprinklerSystem"),
                ("RoofGeometry",     "RoofGeometry"),
                ("FoundationType",   "FoundationType"),
                ("WallSiding",       "WallSiding"),
                ("WallType",         "WallType"),
                ("SoftStory",        "SoftStory"),
            ]
        else:
            norm_pairs = [
                ("YEARBUILT",   "YEARBUILT"),
                ("YEARUPGRAD",  "YEARUPGRAD"),
                ("NUMSTORIES",  "NUMSTORIES"),
                ("NUMBLDGS",    "NUMBLDGS"),
                ("FLOORAREA",   "FLOORAREA"),
                ("EQCV1VAL",    "EQCV1VAL"),
                ("EQCV2VAL",    "EQCV2VAL"),
                ("EQCV3VAL",    "EQCV3VAL"),
                ("EQCV1LCUR",   "EQCV1LCUR"),
                ("SPRINKLER",   "SPRINKLER"),
                ("ROOFGEOM",    "ROOFGEOM"),
                ("FOUNDATION",  "FOUNDATION"),
                ("CLADDING",    "CLADDING"),
                ("WALLTYPE",    "WALLTYPE"),
                ("SOFTSTORY",   "SOFTSTORY"),
            ]

        for canonical, after_col in norm_pairs:
            src = canonical_to_src.get(canonical)
            pairs.append({"label": canonical, "before": src, "after": after_col})

    else:
        raise HTTPException(status_code=400, detail="Invalid step")

    # Derive flat before/after column lists from pairs (deduplicated, preserving order)
    seen_before: set = set()
    before_cols = []
    for p in pairs:
        c = p.get("before")
        if c and c not in seen_before:
            seen_before.add(c)
            before_cols.append(c)
    after_cols = [p["after"] for p in pairs if p.get("after")]

    # Detect full_address_mode from pairs metadata
    full_address_mode = any(p.get("before_is_full_address") for p in pairs)
    full_address_src  = next((p["before"] for p in pairs if p.get("before_is_full_address")), None)

    limit = 100
    rows_data = []

    for i in range(min(len(raw_rows), len(after_rows))):
        before_data = {c: raw_rows[i].get(c) for c in before_cols}
        after_data  = {c: after_rows[i].get(c)  for c in after_cols}
        # Only include row if it has some relevant data on either side
        if any(v is not None and str(v).strip() != "" for v in list(before_data.values()) + list(after_data.values())):
            rows_data.append({"before": before_data, "after": after_data})

    return {
        "step": step,
        "columns": {"before": before_cols, "after": after_cols},
        "pairs": pairs,
        "full_address_mode": full_address_mode,
        "full_address_src":  full_address_src,
        "rows": rows_data[:limit],
        "total": len(rows_data)
    }


# ── Mapping Memory ─────────────────────────────────────────────────────────────────

@app.get("/mapping-memory", tags=["Memory"])
def get_mapping_memory(target_format: Optional[str] = Query(None, regex="^(AIR|RMS)$")):
    """
    List all learned column mapping memories.
    Optionally filter by target_format (AIR or RMS).
    Returns entries sorted by confirmed_count descending.
    """
    entries = mapping_memory.list_memory(target_format)
    stats = mapping_memory.memory_stats()
    return {
        "entries": entries,
        "total": len(entries),
        "stats": stats,
    }


@app.delete("/mapping-memory/{source_col}", tags=["Memory"])
def forget_mapping(source_col: str, target_format: str = Query("AIR", regex="^(AIR|RMS)$")):
    """
    Forget a single learned mapping.
    source_col should be URL-encoded (e.g. 'building%20value').
    target_format must be 'AIR' or 'RMS'.
    """
    deleted = mapping_memory.forget_mapping(source_col, target_format)
    if not deleted:
        raise HTTPException(
            status_code=404,
            detail=f"No memory entry found for '{source_col}' / {target_format}.",
        )
    logger.info(f"Memory forgotten: '{source_col}' / {target_format}")
    return {"forgotten": source_col, "target_format": target_format}


@app.get("/health", tags=["Health"])
def health():
    return {"status": "ok", "version": "1.0.0"}


# ── Internal helpers ───────────────────────────────────────────────────────────

def _apply_column_map(raw_rows: List[Dict], column_map: Dict[str, Optional[str]]) -> List[Dict]:
    """
    Return a new list of row dicts where keys are canonical field names.
    Enforces 1:1 at apply-time: if two source columns map to the same canonical,
    the first one encountered wins and the second is skipped with a warning.
    Source columns mapped to None are dropped (not passed through).
    """
    # Build a deduplicated ordered mapping: canonical → first src col that claimed it
    canonical_claimed_by: Dict[str, str] = {}
    for src_col, canonical in column_map.items():
        if canonical is None:
            continue
        if canonical not in canonical_claimed_by:
            canonical_claimed_by[canonical] = src_col
        else:
            logger.warning(
                f"_apply_column_map: canonical '{canonical}' already claimed by "
                f"'{canonical_claimed_by[canonical]}'; skipping duplicate source '{src_col}'."
            )

    result = []
    for row in raw_rows:
        new_row: Dict[str, Any] = {}
        for canonical, src_col in canonical_claimed_by.items():
            new_row[canonical] = row.get(src_col)
        result.append(new_row)
    return result


def _find_code_columns(column_map: Dict, field_type: str, target: str):
    """
    Identify scheme and value columns for occupancy or construction.
    Returns (scheme_col, value_col) as canonical names.
    """
    # AIR canonical names
    _AIR_OCC_VALUE   = ["OccupancyCode"]
    _AIR_OCC_SCHEME  = ["OccupancyCodeType"]
    _AIR_CONST_VALUE = ["ConstructionCode"]
    _AIR_CONST_SCHEME= ["ConstructionCodeType"]
    # RMS canonical names
    _RMS_OCC_VALUE   = ["OCCTYPE"]
    _RMS_OCC_SCHEME  = ["OCCSCHEME"]
    _RMS_CONST_VALUE = ["BLDGCLASS"]
    _RMS_CONST_SCHEME= ["BLDGSCHEME"]

    if field_type == "occupancy":
        scheme_options = _AIR_OCC_SCHEME  if target == "AIR" else _RMS_OCC_SCHEME
        value_options  = _AIR_OCC_VALUE   if target == "AIR" else _RMS_OCC_VALUE
    else:
        scheme_options = _AIR_CONST_SCHEME if target == "AIR" else _RMS_CONST_SCHEME
        value_options  = _AIR_CONST_VALUE  if target == "AIR" else _RMS_CONST_VALUE

    mapped_vals = set(v for v in column_map.values() if v)
    scheme_col = next((c for c in scheme_options if c in mapped_vals), "")
    value_col  = next((c for c in value_options  if c in mapped_vals), "")
    return scheme_col, value_col


# ── Slip Summary (Done Page Dashboard) ────────────────────────────────────────

def _safe_float(v) -> float:
    """Parse a value to float, stripping currency symbols and commas."""
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "").replace("$", "").replace("£", "").replace("€", "").replace("¥", "").replace("₹", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _bucket_year(year_val) -> str:
    s = str(year_val or "").strip()
    try:
        y = int(float(s))
        if y <= 0:
            return "Unknown"
        if y < 1995:
            return "Pre 1995"
        if y <= 2001:
            return "1995 – 2001"
        if y <= 2010:
            return "2002 – 2010"
        if y <= 2017:
            return "2011 – 2017"
        return "Post 2017"
    except (ValueError, TypeError):
        return "Unknown"


def _bucket_stories(stories_val) -> str:
    s = str(stories_val or "").strip()
    try:
        n = int(float(s))
        if n <= 0:
            return "Unknown"
        if n == 1:
            return "1"
        if n <= 3:
            return "2–3"
        if n <= 7:
            return "4–7"
        return "7+"
    except (ValueError, TypeError):
        return "Unknown"


@app.get("/slip-summary/{upload_id}", tags=["Output"])
def slip_summary(upload_id: str):
    """
    Aggregate final_rows into SlipCoding dashboard metrics:
    location_values, country_state, top_locations,
    occupancy_dist, construction_dist, year_built_dist, stories_dist.
    """
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/slip-summary")

    target = session.get("target_format", "AIR")
    final_rows = session.get("final_rows", [])
    column_map = session.get("column_map", {})

    # Determine value column names based on target
    if target == "AIR":
        bldg_col    = "BuildingValue"
        cont_col    = "ContentsValue"
        bi_col      = "TimeElementValue"
        tiv_col     = "TIV"
        occ_col     = "Occupancy_Description"
        const_col   = "Construction_Description"
        year_col    = "YearBuilt"
        stories_col = "NumberOfStories"
        country_col = "CountryISO"
        state_col   = "Area"
        city_col    = "City"
        street_col  = "Street"
        zip_col     = "PostalCode"
        loc_id_col  = "LocationID"
    else:  # RMS
        bldg_col    = "EQCV1VAL"
        cont_col    = "EQCV2VAL"
        bi_col      = "EQCV3VAL"
        tiv_col     = "TIV"
        occ_col     = "Occupancy_Description"
        const_col   = "Construction_Description"
        year_col    = "YEARBUILT"
        stories_col = "NUMSTORIES"
        country_col = "CNTRYCODE"
        state_col   = "STATECODE"
        city_col    = "CITY"
        street_col  = "STREETNAME"
        zip_col     = "POSTALCODE"
        loc_id_col  = "LOCNUM"

    # ── Compute per-row TIV ───────────────────────────────────────────────────
    total_bldg = total_cont = total_bi = total_tiv_col = 0.0
    country_state_map: Dict[str, Dict] = {}
    loc_rows = []
    occ_map: Dict[str, float] = {}
    const_map: Dict[str, float] = {}
    year_map: Dict[str, float] = {}
    story_map: Dict[str, float] = {}

    for row in final_rows:
        bldg = _safe_float(row.get(bldg_col))
        cont = _safe_float(row.get(cont_col))
        bi   = _safe_float(row.get(bi_col))
        raw_tiv = _safe_float(row.get(tiv_col))

        # Row-level TIV: if a dedicated TIV column exists use it,
        # otherwise sum individual value columns
        if raw_tiv > 0:
            row_tiv = raw_tiv
            # Distribute proportionally for bldg/cont/bi totals only when dedicated TIV is used
            total_bldg  += bldg
            total_cont  += cont
            total_bi    += bi
            total_tiv_col += raw_tiv
        else:
            row_tiv = bldg + cont + bi
            total_bldg += bldg
            total_cont += cont
            total_bi   += bi

        # Country / state
        country = str(row.get(country_col) or "Unknown").strip() or "Unknown"
        state   = str(row.get(state_col) or "NA").strip() or "NA"
        cs_key  = f"{country}||{state}"
        if cs_key not in country_state_map:
            country_state_map[cs_key] = {"country": country, "state": state, "count": 0, "tiv": 0.0}
        country_state_map[cs_key]["count"] += 1
        country_state_map[cs_key]["tiv"]   += row_tiv

        # Top locations (store raw row info + tiv)
        loc_rows.append({
            "loc_id":  str(row.get(loc_id_col) or ""),
            "address": str(row.get(street_col) or ""),
            "city":    str(row.get(city_col) or ""),
            "state":   str(row.get(state_col) or ""),
            "zip":     str(row.get(zip_col) or ""),
            "tiv":     row_tiv,
        })

        # Occupancy
        occ = str(row.get(occ_col) or row.get("OccupancyCode") or "Unknown").strip() or "Unknown"
        occ_map[occ] = occ_map.get(occ, 0.0) + row_tiv

        # Construction
        const = str(row.get(const_col) or row.get("ConstructionCode") or "Unknown").strip() or "Unknown"
        const_map[const] = const_map.get(const, 0.0) + row_tiv

        # Year Built
        yb = _bucket_year(row.get(year_col))
        year_map[yb] = year_map.get(yb, 0.0) + row_tiv

        # Stories
        st = _bucket_stories(row.get(stories_col))
        story_map[st] = story_map.get(st, 0.0) + row_tiv

    grand_total = total_tiv_col if total_tiv_col > 0 else (total_bldg + total_cont + total_bi)

    # Top 10 locations by TIV
    top_locs = sorted(loc_rows, key=lambda r: r["tiv"], reverse=True)[:10]

    # Country/state sorted by TIV desc
    cs_list = sorted(country_state_map.values(), key=lambda r: r["tiv"], reverse=True)

    # Ordered year buckets
    year_order = ["Unknown", "Pre 1995", "1995 – 2001", "2002 – 2010", "2011 – 2017", "Post 2017"]
    year_dist  = [{"label": k, "tiv": year_map.get(k, 0.0)} for k in year_order if k in year_map]

    # Ordered story buckets
    story_order = ["Unknown", "1", "2–3", "4–7", "7+"]
    story_dist  = [{"label": k, "tiv": story_map.get(k, 0.0)} for k in story_order if k in story_map]

    return {
        "total_risks": len(final_rows),
        "location_values": {
            "building": total_bldg,
            "contents": total_cont,
            "bi":       total_bi,
            "total":    grand_total,
        },
        "country_state":      cs_list,
        "top_locations":      top_locs,
        "occupancy_dist":     [{"label": k, "tiv": v} for k, v in sorted(occ_map.items(), key=lambda x: -x[1])],
        "construction_dist":  [{"label": k, "tiv": v} for k, v in sorted(const_map.items(), key=lambda x: -x[1])],
        "year_built_dist":    year_dist,
        "stories_dist":       story_dist,
    }


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
